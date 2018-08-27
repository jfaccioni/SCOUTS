import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# pandas df options
pd.set_option('display.max_rows', 100)
pd.set_option('display.max_columns', 100)
pd.set_option('expand_frame_repr', False)
# cutoff value for gating
GATE_CUTOFF = 0.1


# ## to-do list ## #
# TODO: reformat GUI so that variables below are input by the user

def cytof(input_file, output_folder, outliers, by_marker, tuckey, export_csv,
          export_excel, group_excel, sample_list):

    # get sample names and control sample
    samples = []
    control = ''
    for sample_type, sample in sample_list:
        if sample_type == 'Yes':
            control = sample
        samples.append(sample_list)
    assert control

    # read input as pandas DataFrame, fails if file has unsupported extension
    if input_file.endswith('.xlsx') or input_file.endswith('xls'):
        df = pd.read_excel(input_file)
    elif input_file.endswith('.csv'):
        df = pd.read_csv(input_file)
    else:
        raise IOError

    # gate rows
    for index, row in df.iterrows():
        mean_row_value = np.mean(row[1:])
        if mean_row_value < GATE_CUTOFF:
            df = df.drop(index, axis=0)
    df.reset_index(drop=True, inplace=True)

    # build cutoff values
    marker_dict = {}
    sample_dict = {}  # dict -> { 'sample' : { 'marker' : (. . .) } }
    for sample in sample_list:
        marker_dict = {}  # dict -> { 'marker' : (Q1, Q3, IQR, cutoff value) }
        # select rows containing the string stored in "sample"
        filtered_df = df[df[list(df)[0]].str.contains(sample)]  # don't ask!
        quantile_df = filtered_df.quantile([0.25, 0.75])
        for marker in quantile_df:
            first_quartile, third_quartile = quantile_df[marker]
            iqr = third_quartile - first_quartile
            cutoff = (iqr * tuckey) + third_quartile
            marker_dict[marker] = (first_quartile, third_quartile, iqr, cutoff)
            sample_dict[sample] = marker_dict
    assert sample_dict, marker_dict

    # stream data from pandas to openpyxl
    wb = Workbook()
    raw_data = wb.active
    raw_data.title = 'Gated Raw Data'
    for r in dataframe_to_rows(df, index=False, header=True):
        raw_data.append(r)

    # TODO: CONSERTAR ESTA MERDA

    if by_marker in ['marker', 'both']:

        if outliers in ('control', 'both'):
            for f, n in compare_marker_column(wb, raw_data, sample_dict,
                                              by_control=True, control=control):
                if not group_excel:
                    f.save(os.path.join(output_folder, n))
                    f.remove_sheet(n)

        if outliers in ('sample', 'both'):
            for f, n in compare_marker_column(wb, raw_data, sample_dict,
                                              by_control=False):
                if not group_excel:
                    f.save(os.path.join(output_folder, n))
                    f.remove_sheet(n)

    if by_marker in ['row', 'both']:

        if outliers in ('control', 'both'):
            for f, n in compare_whole_row(wb, raw_data, sample_dict,
                                          by_control=True, control=control):
                if not group_excel:
                    f.save(os.path.join(output_folder, n))
                    f.remove_sheet(n)

        if outliers in ('sample', 'both'):
            for f, n in compare_whole_row(wb, raw_data, sample_dict,
                                          by_control=False):
                if not group_excel:
                    f.save(os.path.join(output_folder, n))
                    f.remove_sheet(n)
    if export_csv:
        pass
    if export_excel and group_excel:
        wb.save(os.path.join(output_folder, 'master_analysis.xlsx'))


def compare_marker_column(xl, data, d, by_control, control=None):
    marker_list = [m for _, (m, (*_)) in d.items()]
    for sample, subdict in d.items():
        query_sample = control
        suf = 'control'
        if not by_control:
            query_sample = sample
            suf = 'sample'
        for marker, values in subdict.items():
            header = True
            row_iterator = 2
            for row in data.iter_rows():
                if header:  # first row
                    header = False
                    continue
                name = row[0].value
                samplename = f'{name}_{marker}_column_outliers_by_{suf}'
                query_position = marker_list.index(marker) + 1
                cell = row[query_position]
                if cell.value > d[query_sample][marker][-1]:  # is outlier
                    if samplename not in xl.sheetnames:
                        row_iterator = 2
                        create_new_sheet(xl, marker, samplename, marker_list)
                        for copy_c in row:
                            xl[samplename].cell(row=row_iterator,
                                                column=copy_c.col_idx,
                                                value=copy_c.value)
                    row_iterator += 1
                yield xl, samplename


def compare_whole_row(xl, data, d, by_control, control=None):
    marker_list = [m for _, (m, (*_)) in d.items()]
    samplename = ''
    for sample, subdict in d.items():
        query_sample = control
        suf = 'control'
        if not by_control:
            query_sample = sample
            suf = 'sample'
        for marker, values in subdict.items():
            header = True
            row_iterator = 2
            for row in data.iter_rows():
                if header:  # first row
                    header = False
                    continue
                for c in row:
                    if c.col_idx == 1:  # first column of rows 2 and beyond
                        name = c.value
                        samplename = f'{name}_{marker}_row_outliers_by_{suf}'
                        continue
                    query_marker = marker_list[c.col_idx - 2]
                    if c.value > d[query_sample][query_marker][-1]:  # outlier
                        if samplename not in xl.sheetnames:
                            row_iterator = 2
                            create_new_sheet(xl, marker, samplename,
                                             marker_list)
                        for copy_c in row:
                            xl[samplename].cell(row=row_iterator,
                                                column=copy_c.col_idx,
                                                value=copy_c.value)
                        row_iterator += 1
                        break
                yield xl, samplename


def create_new_sheet(xl, marker, samplename, marker_list):
    xl.create_sheet(samplename)
    xl[samplename]['A1'] = marker
    for i, header in enumerate(marker_list):  # build header
        xl[samplename].cell(row=1,
                            column=i + 2,
                            value=header)
