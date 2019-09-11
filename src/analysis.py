import os
from collections import namedtuple
from itertools import chain
from typing import Dict, Generator, List, Optional, Tuple

import numpy as np
import pandas as pd
from PySide2.QtWidgets import QMainWindow
from openpyxl import Workbook, load_workbook

from src.utils import NoReferenceError, PandasInputError, SampleNamingError


Stats = namedtuple("Stats", ['first_quartile', 'third_quartile', 'iqr', 'lower_cutoff', 'upper_cutoff'])
Info = namedtuple("Info", ['cutoff_from', 'reference', 'outliers_for', 'category'])


def start_scouts(widget: QMainWindow, input_file: str, output_folder: str, cutoff_rule: str, marker_rule: str,
                 tukey_factor: float, export_csv: bool, export_excel: bool, single_excel: bool,
                 sample_list: List[Tuple[str, str]], gating: str, gate_cutoff_value: Optional[float],
                 export_gated: bool, non_outliers: bool, bottom_outliers: bool) -> None:
    """Main SCOUTS function that organizes user input and calls related functions accordingly."""
    # Loads df and checks for file extension
    df = load_dataframe(input_file=input_file)

    # Sets first column (sample names) as the index
    df.set_index(df.columns[0], inplace=True)

    # Gets markers from dataframe
    markers = get_marker_names(df)

    # Checks if all sample names are in at least one cell of the first column in the df
    samples = get_all_sample_names(sample_list=sample_list)
    validate_sample_names(samples=samples, df=df)

    # Retrieves information on reference, if necessary:
    reference = None
    if 'ref' in cutoff_rule:
        reference = get_reference_sample_name(sample_list=sample_list)

    # Apply gates to df, if any
    if gating == 'cytof':
        apply_cytof_gating(df=df, cutoff=gate_cutoff_value)
    elif gating == 'rnaseq':
        apply_rnaseq_gating(df=df, cutoff=gate_cutoff_value)

    # Gets cutoff dict -> { 'sample' : { 'marker' : Stats(Q1, Q3, IQR, CUTOFF_LOW, CUTOFF_HIGH) } }
    cutoff_df = get_cutoff_dataframe(df=df, samples=samples, markers=markers, reference=reference,
                                     cutoff_rule=cutoff_rule, tukey=tukey_factor)

    # generate outlier tables (SCOUTS)
    run_scouts(widget=widget, df=df, cutoff_df=cutoff_df, samples=samples, markers=markers, reference=reference,
               cutoff_rule=cutoff_rule, marker_rule=marker_rule, export_csv=export_csv, export_excel=export_excel,
               single_excel=single_excel, export_gated=export_gated, non_outliers=non_outliers,
               bottom_outliers=bottom_outliers, output_folder=output_folder)


def load_dataframe(input_file: str) -> pd.DataFrame:
    """Loads input dataframe into memory. Raises an exception if the filename doesn't end with
    .xlsx or .csv (supported formats)."""
    if input_file.endswith('.xlsx') or input_file.endswith('.xls'):
        return pd.read_excel(input_file, header=0)
    elif input_file.endswith('.csv'):
        return pd.read_csv(input_file, header=0)
    else:
        raise PandasInputError


def get_marker_names(df: pd.DataFrame) -> List[str]:
    """Gets the name of all markers from the input DataFrame."""
    return list(df)


def get_all_sample_names(sample_list: List[Tuple[str, str]]) -> List[str]:
    """Gets the name of all samples from the sample table."""
    return [tup[0] for tup in sample_list]


def validate_sample_names(samples: List[str], df: pd.DataFrame) -> None:
    """Checks whether any sample name from the sample table isn't present on the input dataframe.
    Raises an exception if this happens."""
    sample_names = df.index  # Assumes index = sample names (as per documentation)
    for sample in samples:
        if not any(sample in name for name in sample_names):
            raise SampleNamingError
        if sample not in sample_names:
            pass  # TODO: emit warning that sample wasn't found in the DataFrame


def get_reference_sample_name(sample_list: List[Tuple[str, str]]) -> str:
    """Gets the name from the reference sample, raising an exception if it does not exist in the sample table."""
    for sample, sample_type in sample_list:
        if sample_type == 'yes':
            return sample
    else:
        raise NoReferenceError  # If the user chose to run by reference but did not select any reference


def apply_cytof_gating(df: pd.DataFrame, cutoff: float) -> None:
    """Applies gating for Mass Cytometry onto input dataframe, excluding rows with low average expression."""
    indices_to_drop = []
    for index, row in df.iterrows():
        mean_row_value = np.mean(row)
        if mean_row_value <= cutoff:
            indices_to_drop.append(index)
    df.drop(indices_to_drop, axis=0, inplace=True)


# noinspection PyTypeChecker
def apply_rnaseq_gating(df: pd.DataFrame, cutoff: float) -> None:
    """Applies gating for Single-Cell RNASeq onto input dataframe, excluding values below threshold from
    calculations of outlier values."""
    df.mask(df <= cutoff, np.nan, inplace=True)
    df.dropna(how='all', inplace=True)


def get_cutoff_dataframe(df: pd.DataFrame, samples: List[str], markers: List[str], reference: Optional[str],
                         cutoff_rule: str, tukey: float) -> pd.DataFrame:
    """Gets a dataframe with cutoff values(Q1, Q3, IQR, CUTOFF_LOW, CUTOFF_HIGH) in which columns correspond
    to markers and rows (index) correspond to samples."""
    if cutoff_rule == 'ref':
        return get_cutoff(df=df, samples=[reference], markers=markers, tukey=tukey)
    else:
        return get_cutoff(df=df, samples=samples, markers=markers, tukey=tukey)


def get_cutoff(df: pd.DataFrame, samples: List[str], markers: List[str], tukey: float) -> pd.DataFrame:
    """Gets cutoff values for each sample in the list "samples". Returns these values organized as a DataFrame
    where rows represent samples and columns represent markers."""
    result_df = pd.DataFrame(index=samples, columns=markers)
    for sample in samples:
        cutoff_values = get_sample_cutoff(df=df, sample=sample, tukey=tukey)
        result_df.loc[sample] = cutoff_values
    return result_df


def get_sample_cutoff(df: pd.DataFrame, sample: str, tukey: float) -> List[Stats]:
    """Calculates and returns the cutoff statistics for all markers of a given sample."""
    values = []
    filtered_df = filter_df_by_sample_in_index(df=df, sample=sample)
    quantile_df = filtered_df.quantile([0.25, 0.75])
    for marker in quantile_df:
        values.append(get_marker_statistics(tukey=tukey, marker_series=quantile_df[marker]))
    return values


def filter_df_by_sample_in_index(df: pd.DataFrame, sample: str) -> pd.DataFrame:
    """Filters dataframes row according to name passed as argument. Rows whose index contains the sample
    string are selected."""
    return df.loc[df.index.str.contains(sample)]


def get_marker_statistics(tukey: float, marker_series: pd.Series) -> Stats:
    """Calculates and returns the cutoff statistics for a given sample/marker combination (first quartile cutoff,
    third quartile cutoff, interquartile range, lower cutoff and upper cutoff) as a Stats object."""
    first_quartile, third_quartile = marker_series
    iqr = third_quartile - first_quartile
    upper_cutoff = third_quartile + (iqr * tukey)
    lower_cutoff = first_quartile - (iqr * tukey)
    return Stats(first_quartile, third_quartile, iqr, lower_cutoff, upper_cutoff)


def run_scouts(widget: QMainWindow, df: pd.DataFrame, samples: List[str], markers: List[str],
               reference: Optional[str], cutoff_df: pd.DataFrame, cutoff_rule: str, marker_rule: str,
               export_csv: bool, export_excel: bool, single_excel: bool, export_gated: bool, non_outliers: bool,
               bottom_outliers: bool, output_folder: str) -> None:
    """Function responsible for calling SCOUTS subsetting routines, yielding DataFrames, saving them in
    the appropriate format/directory and recording information about each saved result."""
    summary_df = pd.DataFrame(columns=['file number'] + list(Info._fields))
    stats_df_dict = create_stats_dfs(markers=markers, cutoff_rule=cutoff_rule, marker_rule=marker_rule,
                                     samples=samples, bottom=bottom_outliers, non=non_outliers)
    add_whole_population_to_stats_dfs(input_df=df, stats_df_dict=stats_df_dict, samples=samples)
    output_path = os.path.join(output_folder, 'data')
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    excel_file_list = []
    for i, (data, info) in enumerate(yield_dataframes(input_df=df, samples=samples, markers=markers,
                                                      reference=reference, cutoff_df=cutoff_df, cutoff_rule=cutoff_rule,
                                                      marker_rule=marker_rule, non_outliers=non_outliers,
                                                      bottom_outliers=bottom_outliers), 1):
        summary_df = add_scouts_data_to_summary(summary_df, i, info)
        add_scouts_data_to_stats(data, samples, stats_df_dict, info)
        if not widget.stacked_pages.isEnabled():  # user has exited the GUI
            return
        if export_csv:
            csv_path = os.path.join(output_path, '%04d.csv' % i)
            data.to_csv(csv_path)
        if export_excel:
            excel_path = os.path.join(output_path, '%04d.xlsx' % i)
            excel_file_list.append(excel_path)
            data.to_excel(excel_path)
    summary_path = os.path.join(output_folder, 'summary.xlsx')
    generate_summary_table(summary_df, summary_path)
    stats_path = os.path.join(output_folder, 'stats.xlsx')
    generate_stats_table(stats_df_dict, stats_path)
    cutoff_path = os.path.join(output_folder, 'cutoff_values.xlsx')
    generate_cutoff_table(cutoff_df, cutoff_path)
    if export_gated:
        gated_path = os.path.join(output_folder, 'gated_population.xlsx')
        generate_gated_table(df, gated_path)
    if single_excel:
        merged_excel = merge_excel_files(output_path=output_path, summary_path=summary_path, excels=excel_file_list)
        merged_path = os.path.join(output_folder, 'merged_data.xlsx')
        merged_excel.save(merged_path)


def create_stats_dfs(markers: List[str], cutoff_rule: str, marker_rule: str, samples: List[str],
                     bottom: bool, non: bool) -> Dict[str, pd.DataFrame]:
    """str"""  # TODO
    outliers = ['whole population', 'top outliers']
    if non is True:
        outliers += ['non-outliers']
    if bottom is True:
        outliers += ['bottom outliers']
    stats = ['#', 'mean', 'median', 'sd']
    index = pd.MultiIndex.from_product([samples, outliers, stats])
    df_dict = {}
    if 'sample' in cutoff_rule and 'any' in marker_rule:
        df_dict['OutS any marker'] = pd.DataFrame(columns=markers, index=index)
    if 'sample' in cutoff_rule and 'single' in marker_rule:
        df_dict['OutS single marker'] = pd.DataFrame(columns=markers, index=index)
    if 'ref' in cutoff_rule and 'any' in marker_rule:
        df_dict['OutR any marker'] = pd.DataFrame(columns=markers, index=index)
    if 'ref' in cutoff_rule and 'single' in marker_rule:
        df_dict['OutR single marker'] = pd.DataFrame(columns=markers, index=index)
    return df_dict


def add_whole_population_to_stats_dfs(input_df: pd.DataFrame, stats_df_dict: Dict[str, pd.DataFrame],
                                      samples: List[str]) -> None:
    """Adds whole population info (divided by sample) to each stats DataFrame."""
    for stats_df in stats_df_dict.values():
        for sample in samples:
            filtered_df = filter_df_by_sample_in_index(df=input_df, sample=sample)
            values_df = filtered_df.describe().loc[['count', 'mean', '50%', 'std']]
            values_df.index = ['#', 'mean', 'median', 'sd']
            stats_df.loc[sample].loc['whole population'] = values_df.values


def yield_dataframes(input_df: pd.DataFrame, samples: List[str], markers: List[str], reference: Optional[str],
                     cutoff_df: pd.DataFrame, cutoff_rule: str, marker_rule: str, non_outliers: bool,
                     bottom_outliers: bool) -> Generator[pd.DataFrame, None, None]:
    """ Yields dataframes, subsetting input dataframe according to user preferences in the SCOUTS interface."""
    if 'ref' in cutoff_rule:
        if 'any' in marker_rule:
            yield from scouts_by_reference_any_marker(input_df=input_df, cutoff_df=cutoff_df, reference=reference,
                                                      bottom_outliers=bottom_outliers, non_outliers=non_outliers)
        if 'single' in marker_rule:
            yield from scouts_by_reference_single_marker(input_df=input_df, cutoff_df=cutoff_df, markers=markers,
                                                         reference=reference, bottom_outliers=bottom_outliers,
                                                         non_outliers=non_outliers)
    if 'sample' in cutoff_rule:
        if 'any' in marker_rule:
            yield from scouts_by_sample_any_marker(input_df=input_df, cutoff_df=cutoff_df, samples=samples,
                                                   bottom_outliers=bottom_outliers, non_outliers=non_outliers)
        if 'single' in marker_rule:
            yield from scouts_by_sample_single_marker(input_df=input_df, cutoff_df=cutoff_df, samples=samples,
                                                      markers=markers, bottom_outliers=bottom_outliers,
                                                      non_outliers=non_outliers)


# noinspection PyTypeChecker,PyUnresolvedReferences
def scouts_by_reference_any_marker(input_df: pd.DataFrame, cutoff_df: pd.DataFrame, reference: str,
                                   bottom_outliers: bool, non_outliers: bool) -> Generator[pd.DataFrame, None, None]:
    """Subsets DataFrame by reference cutoff, selecting samples that have at least 1 marker above
    outlier cutoff value."""
    upper_cutoffs = [stat.upper_cutoff for stat in cutoff_df.loc[reference]]
    lower_cutoffs = [stat.lower_cutoff for stat in cutoff_df.loc[reference]]
    info_upper = Info('reference', reference, 'any marker', 'top outliers')
    yield input_df.loc[(input_df > upper_cutoffs).any(axis=1)], info_upper
    if bottom_outliers is True:
        info_lower = Info('reference', reference, 'any marker', 'bottom outliers')
        yield input_df.loc[(input_df < lower_cutoffs).any(axis=1)], info_lower
    if non_outliers is True:
        info_non = Info('reference', reference, 'any marker', 'non-outliers')
        yield input_df.loc[(input_df < upper_cutoffs).any(axis=1) &
                           (input_df > lower_cutoffs).any(axis=1)], info_non


def scouts_by_reference_single_marker(input_df: pd.DataFrame, cutoff_df: pd.DataFrame, markers: List[str],
                                      reference: str, bottom_outliers: bool,
                                      non_outliers: bool) -> Generator[pd.DataFrame, None, None]:
    """Subsets DataFrame by reference cutoff, selecting samples that are outliers for each marker (yields each
    dataframe separately)."""
    for marker in markers:
        upper_cutoff = cutoff_df.loc[reference, marker].upper_cutoff
        lower_cutoff = cutoff_df.loc[reference, marker].lower_cutoff
        info_upper = Info('reference', reference, marker, 'top outliers')
        yield input_df.loc[input_df[marker] > upper_cutoff], info_upper
        if bottom_outliers is True:
            info_lower = Info('reference', reference, marker, 'bottom outliers')
            yield input_df.loc[input_df[marker] < lower_cutoff], info_lower
        if non_outliers is True:
            info_non = Info('reference', reference, marker, 'non-outliers')
            yield input_df.loc[(input_df[marker] < upper_cutoff) &
                               (input_df[marker] > lower_cutoff)], info_non


# noinspection PyTypeChecker,PyUnresolvedReferences
def scouts_by_sample_any_marker(input_df: pd.DataFrame, cutoff_df: pd.DataFrame, samples: List[str],
                                bottom_outliers: bool, non_outliers: bool) -> Generator[pd.DataFrame, None, None]:
    """Subsets DataFrame by sample cutoff, selecting samples that have at least 1 marker above
    outlier cutoff value."""
    upper_dfs = []
    non_dfs = []
    lower_dfs = []
    for sample in samples:
        upper_cutoffs = [stat.upper_cutoff for stat in cutoff_df.loc[sample]]
        lower_cutoffs = [stat.lower_cutoff for stat in cutoff_df.loc[sample]]
        filtered_df = filter_df_by_sample_in_index(df=input_df, sample=sample)
        upper_dfs.append(filtered_df.loc[(filtered_df > upper_cutoffs).any(axis=1)])
        if non_outliers is True:
            non_dfs.append(filtered_df.loc[(filtered_df < upper_cutoffs).any(axis=1) &
                                           (filtered_df > lower_cutoffs).any(axis=1)])
        if bottom_outliers is True:
            lower_dfs.append(filtered_df.loc[(filtered_df < lower_cutoffs).any(axis=1)])
    info_upper = Info('sample', 'n/a', 'any marker', 'top outliers')
    yield pd.concat(upper_dfs), info_upper
    if non_outliers is True:
        info_non = Info('sample', 'n/a', 'any marker', 'non-outliers')
        yield pd.concat(non_dfs), info_non
    if bottom_outliers is True:
        info_lower = Info('sample', 'n/a', 'any marker', 'bottom outliers')
        yield pd.concat(lower_dfs), info_lower


def scouts_by_sample_single_marker(input_df: pd.DataFrame, cutoff_df: pd.DataFrame, samples: List[str],
                                   markers: List[str], bottom_outliers: bool,
                                   non_outliers: bool) -> Generator[pd.DataFrame, None, None]:
    """Subsets DataFrame by sample cutoff, selecting samples that are outliers for each marker (yields each
    dataframe separately)."""
    for marker in markers:
        upper_dfs = []
        non_dfs = []
        lower_dfs = []
        for sample in samples:
            upper_cutoff = cutoff_df.loc[sample, marker].upper_cutoff
            lower_cutoff = cutoff_df.loc[sample, marker].lower_cutoff
            filtered_df = filter_df_by_sample_in_index(df=input_df, sample=sample)
            upper_dfs.append(filtered_df.loc[filtered_df[marker] > upper_cutoff])
            if non_outliers is True:
                non_dfs.append(filtered_df.loc[(filtered_df[marker] < upper_cutoff) &
                                               (filtered_df[marker] > lower_cutoff)])
            if bottom_outliers is True:
                lower_dfs.append(filtered_df.loc[filtered_df[marker] < lower_cutoff])

        info_upper = Info('sample', 'n/a', marker, 'top outliers')
        yield pd.concat(upper_dfs), info_upper
        if non_outliers is True:
            info_non = Info('sample', 'n/a', marker, 'non-outliers')
            yield pd.concat(non_dfs), info_non
        if bottom_outliers is True:
            info_lower = Info('sample', 'n/a', marker, 'bottom outliers')
            yield pd.concat(lower_dfs), info_lower


def add_scouts_data_to_summary(df: pd.DataFrame, i: int, info: Info) -> pd.DataFrame:
    """Adds info to the summary df with each new yielded dataframe from SCOUTS."""
    series = pd.Series([i, *info._asdict().values()], index=df.columns, name=f'{len(df)}')
    return df.append(series)


def add_scouts_data_to_stats(data: pd.DataFrame, samples: List[str], stats_df_dict: Dict[str, pd.DataFrame],
                             info: Info) -> None:
    """str"""  # TODO
    for sample in samples:
        values_df = get_values_df(data, sample, info)
        key = get_key_from_info(info)
        df = stats_df_dict[key]
        if 'any' in info.outliers_for:
            df.loc[sample].loc[info.category] = values_df.values
        else:
            df.loc[sample].at[info.category, info.outliers_for] = values_df.values


def get_values_df(data: pd.DataFrame, sample: str, info: Info) -> pd.DataFrame:
    """str"""  # TODO
    filtered_df = filter_df_by_sample_in_index(df=data, sample=sample)
    if 'any' in info.outliers_for:
        values_df = filtered_df.describe().loc[['count', 'mean', '50%', 'std']]
    else:
        values_df = filtered_df.describe().loc[['count', 'mean', '50%', 'std'], info.outliers_for]
    values_df.index = ['#', 'mean', 'median', 'sd']
    return values_df


def get_key_from_info(info: Info) -> str:
    """str"""  # TODO
    if info.cutoff_from == 'sample':
        if 'any marker' == info.outliers_for:
            return 'OutS any marker'
        else:
            return 'OutS single marker'
    else:
        if 'any marker' == info.outliers_for:
            return 'OutR any marker'
        else:
            return 'OutR single marker'


def generate_summary_table(summary_df: pd.DataFrame, summary_path: str) -> None:
    """Generates table with summary of each file generated by SCOUTS and their meaning
    (i.e. how they were generated)."""
    summary_df.to_excel(summary_path, sheet_name='Summary', index=False)


def generate_stats_table(stats_df_dict: Dict[str, pd.DataFrame], stats_path: str) -> None:
    """str"""  # TODO
    writer = pd.ExcelWriter(stats_path)
    for name, df in stats_df_dict.items():
        df.to_excel(writer, sheet_name=name)
    writer.save()


def generate_cutoff_table(cutoff_df: pd.DataFrame, summary_path: str) -> None:
    """Generates table with cutoff values for each sample/marker combination (both high and low)."""
    uppers = [f'{marker}_upper_cutoff' for marker in cutoff_df.columns]
    lowers = [f'{marker}_lower_cutoff' for marker in cutoff_df.columns]
    columns = [m for m in chain.from_iterable(zip(uppers, lowers))]
    output_cutoff_df = pd.DataFrame(index=cutoff_df.index, columns=columns)
    for sample in cutoff_df.index:
        for marker in cutoff_df.columns:
            output_cutoff_df.at[sample, f'{marker}_upper_cutoff'] = cutoff_df.at[sample, marker].upper_cutoff
            output_cutoff_df.at[sample, f'{marker}_lower_cutoff'] = cutoff_df.at[sample, marker].lower_cutoff
    output_cutoff_df.to_excel(summary_path, sheet_name='Cutoff', index_label='Sample')


def generate_gated_table(gated_df: pd.DataFrame, gated_path: str) -> None:
    """Generates table with gated population, i.e. same as the input file except for the gated cells."""
    gated_df.to_excel(gated_path, sheet_name='Gated Population', index=False)


def merge_excel_files(output_path: str, summary_path: str, excels: List[str]) -> Workbook:
    """Merges all Excel files created into a single file and saves it."""
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = 'Summary'
    summary = read_excel_data(summary_path)
    for row in summary:
        wb['Summary'].append(row)
    for index, file in enumerate(excels, 1):
        data = read_excel_data(os.path.join(output_path, file))
        sheet_name = "%04d" % index
        wb.create_sheet(sheet_name)
        for row in data:
            wb[sheet_name].append(row)
    return wb


def read_excel_data(path: str) -> List[List[Optional[str, float]]]:
    """Reads file from the first worksheet from an Excel workbook as a list of list of values."""
    ws = load_workbook(path).active
    rows = []
    for row in ws.iter_rows():
        cells = []
        for cell in row:
            cells.append(cell.value)
        rows.append(cells)
    return rows
